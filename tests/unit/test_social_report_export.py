from __future__ import annotations

import json
from io import BytesIO
from pathlib import Path

from PIL import Image
from openpyxl import load_workbook

from sirius_toolbox.exporters.social_report import export_social_records


class _FakeResponse:
    def __init__(self, content: bytes) -> None:
        self.content = content
        self.headers = {"content-type": "image/png"}

    def raise_for_status(self) -> None:
        return


class _FakeClient:
    def __init__(self, *args, **kwargs) -> None:  # noqa: ANN003
        pass

    def __enter__(self) -> "_FakeClient":
        return self

    def __exit__(self, exc_type, exc, tb) -> None:  # noqa: ANN001
        return

    def get(self, url: str) -> _FakeResponse:
        image = Image.new("RGBA", (12, 12), (255, 0, 0, 128))
        buf = BytesIO()
        image.save(buf, format="PNG")
        return _FakeResponse(buf.getvalue())


def _write_jsonl(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = "\n".join(json.dumps(row, ensure_ascii=False) for row in rows)
    path.write_text(payload + "\n", encoding="utf-8")


def test_export_social_records_task_isolated_and_jpg(monkeypatch, tmp_path: Path) -> None:
    monkeypatch.setattr("sirius_toolbox.exporters.social_report.httpx.Client", _FakeClient)

    input_file = tmp_path / "task_a" / "social_post.jsonl"
    _write_jsonl(
        input_file,
        [
            {
                "source_id": "abc123",
                "task_id": "task-a",
                "images": ["https://example.com/a.png"],
            }
        ],
    )

    out_dir = tmp_path / "exports"
    result = export_social_records(input_file, out_dir, download_images=True)

    task_dir = out_dir / "task-a"
    assert (task_dir / "social_posts.xlsx").exists()
    assert (task_dir / "social_posts.html").exists()
    jpg_files = sorted((task_dir / "images").glob("*.jpg"))
    assert len(jpg_files) == 1
    assert result.downloaded_images == 1


def test_export_social_records_directory_outputs_per_task(tmp_path: Path) -> None:
    input_dir = tmp_path / "tasks"
    _write_jsonl(
        input_dir / "task1" / "social_post.jsonl",
        [{"source_id": "s1", "task_id": "task1", "images": []}],
    )
    _write_jsonl(
        input_dir / "task2" / "social_post.jsonl",
        [{"source_id": "s2", "task_id": "task2", "images": []}],
    )

    out_dir = tmp_path / "exports"
    export_social_records(input_dir, out_dir, download_images=False)

    assert (out_dir / "task1" / "social_posts.xlsx").exists()
    assert (out_dir / "task1" / "social_posts.html").exists()
    assert (out_dir / "task2" / "social_posts.xlsx").exists()
    assert (out_dir / "task2" / "social_posts.html").exists()


def test_export_social_records_publish_time_fallback_and_extra_fields(tmp_path: Path) -> None:
    input_file = tmp_path / "task_x" / "social_post.jsonl"
    _write_jsonl(
        input_file,
        [
            {
                "platform": "xiaohongshu",
                "source_id": "n001",
                "task_id": "task-x",
                "text": "护肤经验分享 #标签 编辑于 2025-12-13",
                "publish_time": "",
                "images": ["https://example.com/a.png"],
                "tags": ["#标签"],
                "url": "https://www.xiaohongshu.com/explore/n001",
                "author_id": "abc123",
                "author_profile_url": "https://www.xiaohongshu.com/user/profile/abc123",
                "note_type": "article",
                "ip_location": "IP属地：上海",
                "like_count": 12000,
                "collect_count": 368,
                "comment_count": 97,
                "share_count": 21,
                "like_count_text": "1.2万",
                "extra_metric": 42,
            }
        ],
    )

    out_dir = tmp_path / "exports"
    export_social_records(input_file, out_dir, download_images=False)

    wb = load_workbook(out_dir / "task-x" / "social_posts.xlsx")
    ws = wb["social_posts"]
    headers = [cell.value for cell in ws[1]]
    values = [cell.value for cell in ws[2]]
    row = dict(zip(headers, values, strict=False))

    assert row["publish_time"] == "2025-12-13"
    assert row["publish_time_source"] == "text_tail"
    assert row["platform"] == "xiaohongshu"
    assert row["author_id"] == "abc123"
    assert row["author_profile_url"] == "https://www.xiaohongshu.com/user/profile/abc123"
    assert row["note_type"] == "article"
    assert row["ip_location"] == "IP属地：上海"
    assert row["like_count"] == 12000
    assert row["collect_count"] == 368
    assert row["comment_count"] == 97
    assert row["share_count"] == 21
    assert row["like_count_text"] == "1.2万"
    assert row["text_length"] == len("护肤经验分享 #标签 编辑于 2025-12-13")
    assert row["tags_count"] == 1
    assert row["image_count"] == 1
    assert row["url_host"] == "www.xiaohongshu.com"
    assert '"extra_metric": 42' in str(row["extra_fields_json"])

    html_text = (out_dir / "task-x" / "social_posts.html").read_text(encoding="utf-8")
    assert "likes=12000" in html_text
    assert "collects=368" in html_text
    assert "comments=97" in html_text
    assert "shares=21" in html_text
    assert "author_profile" in html_text
    assert "ip=IP属地：上海" in html_text
