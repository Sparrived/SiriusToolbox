from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from html import escape
from io import BytesIO
import json
from pathlib import Path
import re
import shutil
from typing import Any, cast
from urllib.parse import urlparse

import httpx
from PIL import Image
from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.worksheet import Worksheet


@dataclass(slots=True)
class ExportResult:
    record_count: int
    downloaded_images: int
    image_dir: Path
    excel_path: Path
    html_path: Path


def _read_jsonl(path: Path, limit: int = 0) -> list[dict[str, Any]]:
    if path.is_dir():
        files = sorted(
            path.glob("*/social_post.jsonl"),
            key=lambda p: p.stat().st_mtime,
        )
        merged: list[dict[str, Any]] = []
        for file_path in files:
            merged.extend(_read_jsonl(file_path, limit=0))
        if limit > 0:
            return merged[-limit:]
        return merged

    if not path.exists():
        return []

    rows: list[dict[str, Any]] = []
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        try:
            row = json.loads(line)
        except json.JSONDecodeError:
            continue
        if isinstance(row, dict):
            rows.append(row)

    if limit > 0:
        return rows[-limit:]
    return rows


def _as_list(value: Any) -> list[str]:
    if not isinstance(value, list):
        return []
    return [str(item) for item in value if item is not None and str(item).strip()]


def _clean_text(value: Any) -> str:
    return str(value or "").strip()


def _first_non_empty(record: dict[str, Any], keys: list[str]) -> str:
    for key in keys:
        value = _clean_text(record.get(key))
        if value:
            return value
    return ""


def _extract_publish_time_from_text(text: str) -> str:
    clean = _clean_text(text)
    if not clean:
        return ""

    patterns = [
        r"(?:编辑于|发布于)\s*(\d{4}[-/.]\d{1,2}[-/.]\d{1,2}(?:\s+\d{1,2}:\d{2}(?::\d{2})?)?)",
        r"(\d{4}[-/.]\d{1,2}[-/.]\d{1,2}(?:\s+\d{1,2}:\d{2}(?::\d{2})?)?)$",
        r"(\d{1,2}[-/.]\d{1,2}(?:\s+\d{1,2}:\d{2}(?::\d{2})?)?)$",
    ]
    for pattern in patterns:
        match = re.search(pattern, clean)
        if match:
            return _clean_text(match.group(1))
    return ""


def _extract_url_host(url: str) -> str:
    try:
        return urlparse(url).netloc or ""
    except Exception:  # noqa: BLE001
        return ""


def _guess_title(record: dict[str, Any]) -> str:
    title = _clean_text(record.get("title"))
    if title:
        return title

    text = _clean_text(record.get("text"))
    if not text:
        return ""

    first_line = text.splitlines()[0].strip()
    if len(first_line) <= 40:
        return first_line
    return first_line[:40] + "..."


def _normalize_record(record: dict[str, Any]) -> dict[str, Any]:
    normalized = dict(record)
    publish_time = _first_non_empty(
        normalized,
        [
            "publish_time",
            "publishTime",
            "published_at",
            "publishedAt",
            "create_time",
            "createTime",
            "note_time",
            "noteTime",
            "time",
            "date",
            "发布时间",
            "发布于",
        ],
    )
    publish_time_source = "field"
    if not publish_time:
        publish_time = _extract_publish_time_from_text(_clean_text(normalized.get("text")))
        publish_time_source = "text_tail" if publish_time else ""

    normalized["source_id"] = _clean_text(normalized.get("source_id"))
    normalized["platform"] = _clean_text(normalized.get("platform"))
    normalized["author"] = _clean_text(normalized.get("author"))
    normalized["title"] = _guess_title(normalized)
    normalized["text"] = _clean_text(normalized.get("text"))
    normalized["publish_time"] = publish_time
    normalized["publish_time_source"] = publish_time_source
    normalized["url"] = _clean_text(normalized.get("url"))
    normalized["url_host"] = _extract_url_host(normalized["url"])
    normalized["note_type"] = _clean_text(normalized.get("note_type"))
    normalized["author_profile_url"] = _clean_text(normalized.get("author_profile_url"))
    normalized["author_id"] = _clean_text(normalized.get("author_id"))
    normalized["ip_location"] = _clean_text(normalized.get("ip_location"))
    normalized["like_count"] = int(normalized.get("like_count") or 0)
    normalized["collect_count"] = int(normalized.get("collect_count") or 0)
    normalized["comment_count"] = int(normalized.get("comment_count") or 0)
    normalized["share_count"] = int(normalized.get("share_count") or 0)
    normalized["like_count_text"] = _clean_text(normalized.get("like_count_text"))
    normalized["collect_count_text"] = _clean_text(normalized.get("collect_count_text"))
    normalized["comment_count_text"] = _clean_text(normalized.get("comment_count_text"))
    normalized["share_count_text"] = _clean_text(normalized.get("share_count_text"))
    normalized["tags"] = _as_list(normalized.get("tags"))
    normalized["images"] = _as_list(normalized.get("images"))
    normalized["local_images"] = _as_list(normalized.get("local_images"))
    normalized["collected_at"] = _clean_text(normalized.get("collected_at"))
    normalized["task_id"] = _clean_text(normalized.get("task_id"))
    normalized["raw_ref"] = _clean_text(normalized.get("raw_ref"))
    normalized["local_post_path"] = _clean_text(normalized.get("local_post_path"))
    normalized["text_length"] = len(normalized["text"])
    normalized["tags_count"] = len(normalized["tags"])
    normalized["image_count"] = len(normalized["images"])
    normalized["local_image_count"] = len(normalized["local_images"])

    known_keys = {
        "platform",
        "source_id",
        "title",
        "text",
        "author",
        "publish_time",
        "publishTime",
        "published_at",
        "publishedAt",
        "create_time",
        "createTime",
        "note_time",
        "noteTime",
        "time",
        "date",
        "发布时间",
        "发布于",
        "images",
        "tags",
        "url",
        "collected_at",
        "raw_ref",
        "task_id",
        "local_post_path",
        "local_images",
        "note_type",
        "author_profile_url",
        "author_id",
        "ip_location",
        "like_count",
        "collect_count",
        "comment_count",
        "share_count",
        "like_count_text",
        "collect_count_text",
        "comment_count_text",
        "share_count_text",
    }
    extras = {k: v for k, v in record.items() if k not in known_keys}
    normalized["extra_fields_json"] = json.dumps(extras, ensure_ascii=False, sort_keys=True) if extras else ""
    return normalized


def _safe_stem(text: str) -> str:
    compact = re.sub(r"[^a-zA-Z0-9_-]+", "_", text)
    compact = compact.strip("_")
    return compact[:80] if compact else "item"


def _write_as_jpg(content: bytes, target: Path) -> bool:
    try:
        with Image.open(BytesIO(content)) as image:
            # Keep transparent sources readable when flattening to JPEG.
            if image.mode in {"RGBA", "LA"}:
                background = Image.new("RGB", image.size, (255, 255, 255))
                alpha = image.split()[-1]
                background.paste(image, mask=alpha)
                image = background
            else:
                image = image.convert("RGB")
            image.save(target, format="JPEG", quality=92, optimize=True)
            return True
    except Exception:  # noqa: BLE001
        return False


def _download_images(records: list[dict[str, Any]], image_dir: Path) -> int:
    image_dir.mkdir(parents=True, exist_ok=True)
    downloaded = 0
    cache: dict[str, str] = {}

    with httpx.Client(timeout=20.0, follow_redirects=True) as client:
        for idx, record in enumerate(records):
            source_id = _safe_stem(_clean_text(record.get("source_id")) or f"record_{idx + 1}")
            local_images: list[str] = []
            for img_index, url in enumerate(_as_list(record.get("images")), start=1):
                if url in cache:
                    local_images.append(cache[url])
                    continue

                try:
                    response = client.get(url)
                    response.raise_for_status()
                except Exception:  # noqa: BLE001
                    continue

                filename = f"{source_id}_{img_index:02d}.jpg"
                target = image_dir / filename
                duplicate = 1
                while target.exists():
                    filename = f"{source_id}_{img_index:02d}_{duplicate}.jpg"
                    target = image_dir / filename
                    duplicate += 1

                if not _write_as_jpg(response.content, target):
                    continue

                rel_path = f"images/{filename}"
                cache[url] = rel_path
                local_images.append(rel_path)
                downloaded += 1

            record["local_images"] = local_images

    return downloaded


def _write_excel(records: list[dict[str, Any]], out_path: Path) -> None:
    wb = Workbook()
    ws = cast(Worksheet, wb.active)
    ws.title = "social_posts"

    headers = [
        "platform",
        "task_id",
        "source_id",
        "author",
        "author_id",
        "author_profile_url",
        "title",
        "text",
        "text_length",
        "publish_time",
        "publish_time_source",
        "collected_at",
        "note_type",
        "ip_location",
        "url",
        "url_host",
        "like_count",
        "collect_count",
        "comment_count",
        "share_count",
        "like_count_text",
        "collect_count_text",
        "comment_count_text",
        "share_count_text",
        "image_count",
        "local_image_count",
        "image_urls",
        "local_image_files",
        "tags",
        "tags_count",
        "raw_ref",
        "local_post_path",
        "extra_fields_json",
    ]
    ws.append(headers)

    for record in records:
        images = _as_list(record.get("images"))
        local_images = _as_list(record.get("local_images"))
        tags = _as_list(record.get("tags"))
        ws.append(
            [
                str(record.get("platform") or ""),
                str(record.get("task_id") or ""),
                str(record.get("source_id") or ""),
                str(record.get("author") or ""),
                str(record.get("author_id") or ""),
                str(record.get("author_profile_url") or ""),
                str(record.get("title") or ""),
                str(record.get("text") or ""),
                int(record.get("text_length") or 0),
                str(record.get("publish_time") or ""),
                str(record.get("publish_time_source") or ""),
                str(record.get("collected_at") or ""),
                str(record.get("note_type") or ""),
                str(record.get("ip_location") or ""),
                str(record.get("url") or ""),
                str(record.get("url_host") or ""),
                int(record.get("like_count") or 0),
                int(record.get("collect_count") or 0),
                int(record.get("comment_count") or 0),
                int(record.get("share_count") or 0),
                str(record.get("like_count_text") or ""),
                str(record.get("collect_count_text") or ""),
                str(record.get("comment_count_text") or ""),
                str(record.get("share_count_text") or ""),
                int(record.get("image_count") or len(images)),
                int(record.get("local_image_count") or len(local_images)),
                "\n".join(images),
                "\n".join(local_images),
                ", ".join(tags),
                int(record.get("tags_count") or len(tags)),
                str(record.get("raw_ref") or ""),
                str(record.get("local_post_path") or ""),
                str(record.get("extra_fields_json") or ""),
            ]
        )

    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True)

    widths = {
        "A": 14,
        "B": 38,
        "C": 24,
        "D": 18,
        "E": 24,
        "F": 44,
        "G": 30,
        "H": 62,
        "I": 12,
        "J": 20,
        "K": 16,
        "L": 25,
        "M": 16,
        "N": 24,
        "O": 52,
        "P": 24,
        "Q": 12,
        "R": 12,
        "S": 12,
        "T": 12,
        "U": 20,
        "V": 20,
        "W": 20,
        "X": 20,
        "Y": 12,
        "Z": 14,
        "AA": 72,
        "AB": 44,
        "AC": 36,
        "AD": 12,
        "AE": 44,
        "AF": 42,
        "AG": 52,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[7].alignment = Alignment(wrap_text=True, vertical="top")
        row[25].alignment = Alignment(wrap_text=True, vertical="top")
        row[26].alignment = Alignment(wrap_text=True, vertical="top")
        row[27].alignment = Alignment(wrap_text=True, vertical="top")
        row[30].alignment = Alignment(wrap_text=True, vertical="top")
        row[31].alignment = Alignment(wrap_text=True, vertical="top")
        url_cell = cast(Cell, row[14])
        url_value = url_cell.value
        url_cell.hyperlink = (
            Hyperlink(ref=url_cell.coordinate, target=url_value)
            if isinstance(url_value, str) and url_value
            else None
        )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def _build_card(record: dict[str, Any]) -> str:
    title = escape(str(record.get("title") or "(no title)"))
    text = escape(str(record.get("text") or "").strip())
    author = escape(str(record.get("author") or ""))
    author_id = escape(str(record.get("author_id") or ""))
    author_profile_url = escape(str(record.get("author_profile_url") or ""))
    note_type = escape(str(record.get("note_type") or ""))
    ip_location = escape(str(record.get("ip_location") or ""))
    url = escape(str(record.get("url") or ""))
    publish_time = escape(str(record.get("publish_time") or ""))
    collected_at = escape(str(record.get("collected_at") or ""))
    like_count = int(record.get("like_count") or 0)
    collect_count = int(record.get("collect_count") or 0)
    comment_count = int(record.get("comment_count") or 0)
    share_count = int(record.get("share_count") or 0)
    local_images = _as_list(record.get("local_images"))
    images = local_images if local_images else _as_list(record.get("images"))

    image_html = ""
    for img in images[:4]:
        safe = escape(img)
        image_html += f'<a href="{safe}" target="_blank" rel="noreferrer"><img src="{safe}" loading="lazy" /></a>'

    author_html = f"author={author}"
    if author_id:
        author_html += f" ({author_id})"
    if author_profile_url:
        author_html += f" | <a href='{author_profile_url}' target='_blank' rel='noreferrer'>author_profile</a>"

    detail_parts = []
    if note_type:
        detail_parts.append(f"type={note_type}")
    if ip_location:
        detail_parts.append(f"ip={ip_location}")
    detail_meta = " | ".join(detail_parts)

    return (
        "<article class='card'>"
        f"<h3>{title}</h3>"
        f"<p class='meta'>{author_html}</p>"
        f"<p class='meta'>publish={publish_time} | collected={collected_at}</p>"
        f"<p class='stats'>likes={like_count} | collects={collect_count} | comments={comment_count} | shares={share_count}</p>"
        f"<p class='meta'>{detail_meta}</p>"
        f"<p class='text'>{text or '(no text)'}</p>"
        f"<p><a href='{url}' target='_blank' rel='noreferrer'>{url}</a></p>"
        f"<div class='images'>{image_html}</div>"
        "</article>"
    )


def _write_html(records: list[dict[str, Any]], out_path: Path, downloaded_images: int) -> None:
    cards = "".join(_build_card(record) for record in records)
    generated = datetime.now().isoformat(timespec="seconds")
    html = f"""<!doctype html>
<html lang=\"en\">
<head>
  <meta charset=\"utf-8\" />
  <meta name=\"viewport\" content=\"width=device-width, initial-scale=1\" />
  <title>Social Report</title>
  <style>
    body {{
      margin: 0;
      font-family: Segoe UI, Arial, sans-serif;
      background: #f4f7fb;
      color: #1f2d3d;
    }}
    .wrap {{
      max-width: 1200px;
      margin: 24px auto;
      padding: 0 16px 24px;
    }}
    .head {{
      background: #fff;
      border: 1px solid #d7e2ef;
      border-radius: 12px;
      padding: 14px 16px;
      margin-bottom: 16px;
    }}
    .grid {{
      display: grid;
      grid-template-columns: repeat(auto-fill, minmax(320px, 1fr));
      gap: 14px;
    }}
    .card {{
      background: #fff;
      border: 1px solid #d7e2ef;
      border-radius: 12px;
      padding: 12px;
    }}
    .card h3 {{
      margin: 0 0 8px;
      font-size: 16px;
      line-height: 1.3;
    }}
    .meta {{
      margin: 0 0 8px;
      font-size: 12px;
      color: #506680;
    }}
        .stats {{
            margin: 0 0 8px;
            font-size: 12px;
            color: #2b4b67;
            font-weight: 600;
        }}
    .text {{
      white-space: pre-wrap;
      font-size: 13px;
      max-height: 180px;
      overflow: auto;
      border: 1px solid #e3ebf5;
      border-radius: 8px;
      padding: 8px;
      background: #f9fbfe;
    }}
    .images {{
      margin-top: 8px;
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 8px;
    }}
    .images img {{
      width: 100%;
      height: 160px;
      object-fit: cover;
      border-radius: 8px;
      border: 1px solid #dce7f2;
      background: #eef4fb;
    }}
  </style>
</head>
<body>
  <div class=\"wrap\">
    <section class=\"head\">
      <h1>Social Posts Report</h1>
      <p>records={len(records)} | downloaded_images={downloaded_images} | generated_at={escape(generated)}</p>
    </section>
    <section class=\"grid\">{cards}</section>
  </div>
</body>
</html>
"""

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(html, encoding="utf-8")


def _group_by_task(records: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for record in records:
        task_id = _safe_stem(_clean_text(record.get("task_id")) or "legacy")
        grouped.setdefault(task_id, []).append(record)
    return grouped


def export_social_records(
    input_path: Path,
    output_dir: Path,
    limit: int = 0,
    download_images: bool = True,
) -> ExportResult:
    raw_records = _read_jsonl(input_path, limit=limit)
    records = [_normalize_record(record) for record in raw_records]
    output_dir.mkdir(parents=True, exist_ok=True)
    grouped_records = _group_by_task(records)

    downloaded = 0
    image_dir = output_dir / "images"
    excel_path = output_dir / "social_posts.xlsx"
    html_path = output_dir / "social_posts.html"

    for task_id, task_records in grouped_records.items():
        task_output_dir = output_dir / task_id
        if task_output_dir.exists():
            shutil.rmtree(task_output_dir)
        task_output_dir.mkdir(parents=True, exist_ok=True)

        task_image_dir = task_output_dir / "images"
        task_excel_path = task_output_dir / "social_posts.xlsx"
        task_html_path = task_output_dir / "social_posts.html"

        task_downloaded = 0
        if download_images:
            task_downloaded = _download_images(task_records, task_image_dir)
        _write_excel(task_records, task_excel_path)
        _write_html(task_records, task_html_path, downloaded_images=task_downloaded)

        downloaded += task_downloaded
        image_dir = task_image_dir
        excel_path = task_excel_path
        html_path = task_html_path

    return ExportResult(
        record_count=len(records),
        downloaded_images=downloaded,
        image_dir=image_dir,
        excel_path=excel_path,
        html_path=html_path,
    )
